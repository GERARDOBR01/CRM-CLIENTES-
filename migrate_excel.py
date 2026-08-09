"""Migracion del tracker en Excel (`leads_tracker.xlsx`, pestaña "Leads") a SQLite.

Uso:
    python migrate_excel.py                          # busca el xlsx en rutas comunes
    python migrate_excel.py --excel ruta\\al\\archivo.xlsx
    python migrate_excel.py --reset                  # borra la tabla leads y reimporta
    python migrate_excel.py --actualizar             # sobrescribe leads ya existentes

Se identifica cada lead por el nombre del negocio, asi que correr el script dos
veces no duplica registros.
"""

from __future__ import annotations

import argparse
import sys
import unicodedata
from pathlib import Path

import openpyxl

import db
import whatsapp

RUTAS_CANDIDATAS = [
    Path(__file__).resolve().parent / "leads_tracker.xlsx",
    Path.home() / "Downloads" / "leads_tracker.xlsx",
    Path.home() / "Desktop" / "leads_tracker.xlsx",
    Path.home() / "OneDrive" / "leads_tracker.xlsx",
]


def _norm(texto: str) -> str:
    """minusculas, sin acentos, sin espacios ni signos — para comparar encabezados."""
    sin_acentos = "".join(
        c for c in unicodedata.normalize("NFD", str(texto)) if unicodedata.category(c) != "Mn"
    )
    return "".join(c for c in sin_acentos.lower() if c.isalnum())


# Encabezado normalizado (por prefijo/substring) -> campo del modelo.
REGLAS_COLUMNAS = [
    ("negocio", "negocio"),
    ("categoria", "categoria"),
    ("direccion", "direccion"),
    ("telefono", "telefono"),
    ("plataforma", "plataforma"),
    ("evidencia", "evidencia_dolor"),
    ("mensaje", "mensaje_plantilla"),
    ("estatus", "estatus"),
    ("fechadecontacto", "fecha_contacto"),
    ("fechacontacto", "fecha_contacto"),
    ("proximaaccion", "proxima_accion"),
    ("notas", "notas"),
]


def mapear_columnas(encabezados: list) -> dict[int, str]:
    mapa: dict[int, str] = {}
    for i, enc in enumerate(encabezados):
        if enc is None:
            continue
        clave = _norm(enc)
        for patron, campo in REGLAS_COLUMNAS:
            if patron in clave and campo not in mapa.values():
                mapa[i] = campo
                break
    return mapa


def normalizar_plataforma(valor: str) -> tuple[str, str]:
    """Devuelve (plataforma_canonica, texto_extra_a_conservar).

    El Excel trae valores como "WhatsApp (y explorar LinkedIn del gerente si aparece)";
    guardamos el canonico en el campo y el matiz completo en notas para no perderlo.
    """
    texto = str(valor or "").strip()
    if not texto:
        return "WhatsApp", ""
    clave = _norm(texto)
    canonica = "Otro"
    for p in db.PLATAFORMAS:
        if _norm(p) in clave:
            canonica = p
            break
    extra = texto if _norm(canonica) != clave else ""
    return canonica, extra


def normalizar_estatus(valor) -> str:
    texto = str(valor or "").strip()
    if not texto:
        return "Sin contactar"
    for e in db.ESTATUS:
        if _norm(e) == _norm(texto):
            return e
    return "Sin contactar"


def leer_excel(ruta: Path, hoja: str = "Leads") -> list[dict]:
    wb = openpyxl.load_workbook(ruta, data_only=True)
    if hoja not in wb.sheetnames:
        candidatas = [s for s in wb.sheetnames if _norm(s) == _norm(hoja)]
        if not candidatas:
            raise SystemExit(
                f"No encontré la pestaña '{hoja}' en {ruta.name}. Pestañas: {wb.sheetnames}"
            )
        hoja = candidatas[0]
    ws = wb[hoja]

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return []

    mapa = mapear_columnas(list(filas[0]))
    if "negocio" not in mapa.values():
        raise SystemExit(f"La pestaña '{hoja}' no tiene una columna de Negocio reconocible.")

    leads: list[dict] = []
    for fila in filas[1:]:
        registro = {campo: fila[i] for i, campo in mapa.items() if i < len(fila)}
        negocio = str(registro.get("negocio") or "").strip()
        if not negocio:
            continue

        plataforma, extra = normalizar_plataforma(registro.get("plataforma"))
        telefono_raw = str(registro.get("telefono") or "").strip()

        notas = [str(registro.get("notas") or "").strip()]
        if extra:
            notas.append(f"Plataforma (nota del Excel): {extra}")

        # Si la celda de telefono trae una instruccion en vez de un numero,
        # se conserva como nota y el campo telefono queda vacio.
        if telefono_raw and not whatsapp.es_telefono_valido(telefono_raw):
            notas.append(f"Teléfono (nota del Excel): {telefono_raw}")
            telefono_raw = ""

        leads.append(
            {
                "negocio": negocio,
                "categoria": str(registro.get("categoria") or "").strip(),
                "direccion": str(registro.get("direccion") or "").strip(),
                "telefono": telefono_raw,
                "plataforma": plataforma,
                "evidencia_dolor": str(registro.get("evidencia_dolor") or "").strip(),
                "mensaje_plantilla": str(registro.get("mensaje_plantilla") or "").strip(),
                "estatus": normalizar_estatus(registro.get("estatus")),
                "fecha_contacto": db.normalizar_fecha(registro.get("fecha_contacto")),
                "proxima_accion": str(registro.get("proxima_accion") or "").strip(),
                "notas": "\n".join(n for n in notas if n),
            }
        )
    return leads


def importar(ruta: Path, reset: bool = False, actualizar: bool = False) -> dict:
    db.init_db()
    if reset:
        with db.conectar() as con:
            con.execute("DELETE FROM contactos")
            con.execute("DELETE FROM leads")
            con.execute("DELETE FROM sqlite_sequence WHERE name IN ('leads','contactos')")

    leads = leer_excel(ruta)

    with db.conectar() as con:
        existentes = {
            str(f["negocio"]).strip().lower(): f["id"]
            for f in con.execute("SELECT id, negocio FROM leads").fetchall()
        }

    creados, actualizados, omitidos = 0, 0, 0
    for lead in leads:
        clave = lead["negocio"].lower()
        if clave in existentes:
            if actualizar:
                db.actualizar_lead(existentes[clave], **lead)
                actualizados += 1
            else:
                omitidos += 1
            continue
        db.crear_lead(**lead)
        creados += 1

    return {"leidos": len(leads), "creados": creados, "actualizados": actualizados, "omitidos": omitidos}


def localizar_excel(ruta: str | None) -> Path:
    if ruta:
        p = Path(ruta).expanduser()
        if not p.exists():
            raise SystemExit(f"No existe el archivo: {p}")
        return p
    for p in RUTAS_CANDIDATAS:
        if p.exists():
            return p
    raise SystemExit(
        "No encontré leads_tracker.xlsx. Pásalo con --excel ruta\\al\\archivo.xlsx\n"
        "Rutas revisadas:\n  " + "\n  ".join(str(p) for p in RUTAS_CANDIDATAS)
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Importa leads_tracker.xlsx a leads.db")
    parser.add_argument("--excel", help="Ruta al archivo .xlsx")
    parser.add_argument("--hoja", default="Leads", help="Pestaña a leer (default: Leads)")
    parser.add_argument("--reset", action="store_true", help="Vacía la base antes de importar")
    parser.add_argument(
        "--actualizar", action="store_true", help="Sobrescribe leads que ya existen (por nombre)"
    )
    args = parser.parse_args(argv)

    ruta = localizar_excel(args.excel)
    print(f"Leyendo {ruta}")
    res = importar(ruta, reset=args.reset, actualizar=args.actualizar)
    print(
        f"Leads en el Excel: {res['leidos']} | nuevos: {res['creados']} | "
        f"actualizados: {res['actualizados']} | ya existían: {res['omitidos']}"
    )
    print(f"Base de datos: {db.DB_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
